#!/usr/bin/env python3
"""
convert_ovary_reference.py - build sample_summary's ovary reference assets.

Converts the human ovary supplementary tables from Sci Adv adm7506 into the two
forms sample_summary consumes, and writes them into assets/.

  Data S6 -> assets/ovary_reference_major.csv.gz
      Genes x cell type (Endothelial, Immune, Pericyte, Stromal), mean expression,
      averaged over the three donor sheets. This is the form cluster centroids are
      Spearman-correlated against.

  Data S7 -> assets/ovary_reference_immune.csv.gz
      Genes x immune subtype (NK, T, Macrophage, Mast), mean expression, from the one
      sheet Data S7 carries. Cluster and per-cell profiles are correlated against its
      columns the same way they are against the major reference.

  Data S5 -> assets/ovary_follicle_markers.yaml
      Granulosa / Oocyte / Theca markers as {gene: quality score}. These sheets carry
      per-sample values rather than per-cell-type means, so they cannot be made into
      reference columns and are scored instead. EVERY gene is written out with the
      paper's own 1-5 quality score; which of them to actually use is the deck's
      decision, not this script's.

Run outside the pipeline container, which has no openpyxl — the committed assets are
the gzipped CSV and the YAML, not the workbooks. Mirrors sammy_r21's
scripts/convert_atlas.py.

Usage:
    scripts/convert_ovary_reference.py --supdir ~/Downloads/adm7506 --outdir assets
    scripts/convert_ovary_reference.py --supdir ... --outdir ... --panel <a create_sdata zarr>
"""

import argparse
import datetime
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import yaml

# The three donor sheets in Data S6. Each is genes x (4 mean-expression columns, then
# 4 detection-rate columns) under a two-row header.
S6_DONOR_SHEETS = ["Donor 3", "Donor 4", "Donor 5"]
N_MAJOR_TYPES = 4

# Data S7's single sheet: the same two-row header as an S6 donor sheet, four immune
# subtypes (NK, T, Macrophage, Mast) under "Mean Expression" then "Detection Rate".
S7_SHEET = "immu4_mean_detect"
N_IMMUNE_TYPES = 4

# Excel reads MARCH*/SEPT*/DEC* gene symbols as dates. Month -> the symbol prefix it ate.
EXCEL_DATE_GENES = {3: "MARCH", 9: "SEPT", 12: "DEC"}

# Data S5's follicle sheets. Each lists its gene set in column A, terminated by a blank
# row and followed by a footnote describing the paper's 1-5 marker-quality scoring.
S5_SETS = {
    "granulosa": "Granulosa - 96 Genes",
    "oocyte": "Oocyte - 76 Genes",
    "theca": "Theca - 46 Genes",
}


def parse_args():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--supdir", required=True,
                        help="Directory holding adm7506_Data_S5.xlsx and _S6.xlsx")
    parser.add_argument("--outdir", default="assets", help="Where to write the assets")
    parser.add_argument("--panel",
                        help="Optional create_sdata zarr; reports gene overlap with the "
                             "panel, which is what decides whether the correlation is "
                             "worth trusting")
    return parser.parse_args()


def gene_symbol(value):
    """A cell from a gene column, as a symbol — undoing Excel's date autoformatting.

    MARCH1, SEPT1 and DEC1 are read by Excel as dates and stored as datetimes, so they
    arrive here as `2023-03-01 00:00:00` rather than `MARCH1`. Ten such entries are in
    Data S6 and twelve in Data S7, all of them MARCH family. Left alone they become
    literal date strings in the asset that match no gene anywhere and silently drop out
    of every intersection.

    It costs nothing today — none of the MARCH family is on the 5K panel under either
    the old name or the current MARCHF one, so the corrupted rows never entered a
    correlation. It is fixed because a panel that did carry them would lose them without
    any error, and because the immune reference this now also writes is one where the
    family is relevant (MARCHF1 regulates MHC-II in antigen-presenting cells).

    Note the recovered symbol is the OLD one — MARCH1, not MARCHF1 — because that is
    what the paper's table said before Excel ate it. Renaming to current HGNC would be a
    second, separate decision.
    """
    if isinstance(value, datetime.datetime):
        return f"{EXCEL_DATE_GENES.get(value.month, value.month)}{value.day}"
    return None if value is None else str(value).strip()


def read_immune_reference(path):
    """Genes x immune subtype mean expression, from the single Data S7 sheet.

    The same shape as one of Data S6's donor sheets — a two-row header spanning
    "Mean Expression" then "Detection Rate", four subtypes under each — but there is one
    sheet rather than three, so nothing is averaged. Only the mean-expression block is
    taken, for the same reason as the major reference: detection rate is a different
    quantity and is not comparable to a cluster centroid.

    Worth knowing before this is used to name anything: the four subtype profiles are
    NOT four separable things on a 5K panel. Measured over the panel intersection, the
    NK and T centroids correlate at 0.924 — they cannot be told apart at any depth by any
    metric — while Macrophage sits ~0.69-0.71 from the others and Mast ~0.60.
    """
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook[S7_SHEET].iter_rows(values_only=True))
    workbook.close()
    types = list(rows[1][1:1 + N_IMMUNE_TYPES])
    data = {}
    for row in rows[2:]:
        gene, values = gene_symbol(row[0]), row[1:1 + N_IMMUNE_TYPES]
        if not gene or any(v is None for v in values):
            continue
        data[gene] = [float(v) for v in values]
    reference = pd.DataFrame.from_dict(data, orient="index", columns=types)
    reference.index.name = "gene"
    print(f"  {S7_SHEET}: {reference.shape[0]:,} genes x {reference.shape[1]} subtypes "
          f"({', '.join(types)})")
    return reference


def read_major_reference(path):
    """Genes x major cell type mean expression, averaged over the donor sheets.

    Only the mean-expression block is taken. Detection rate is a different quantity
    and would not be comparable to a cluster centroid.
    """
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    frames = {}
    for sheet in S6_DONOR_SHEETS:
        rows = list(workbook[sheet].iter_rows(values_only=True))
        # Row 0 spans "Mean Expression" / "Detection Rate"; row 1 names the types.
        types = list(rows[1][1:1 + N_MAJOR_TYPES])
        data = {}
        for row in rows[2:]:
            gene, values = gene_symbol(row[0]), row[1:1 + N_MAJOR_TYPES]
            if not gene or any(v is None for v in values):
                continue
            data[gene] = [float(v) for v in values]
        frames[sheet] = pd.DataFrame.from_dict(data, orient="index", columns=types)
        print(f"  {sheet}: {frames[sheet].shape[0]:,} genes x {len(types)} types")
    workbook.close()

    # Intersect before averaging, so no column mean is taken over a different set of
    # donors for one gene than for another.
    common = sorted(set.intersection(*(set(f.index) for f in frames.values())))
    stacked = np.stack([frames[s].loc[common].to_numpy() for s in S6_DONOR_SHEETS])
    reference = pd.DataFrame(stacked.mean(axis=0), index=common,
                             columns=frames[S6_DONOR_SHEETS[0]].columns)
    reference.index.name = "gene"
    print(f"  averaged over {len(S6_DONOR_SHEETS)} donors "
          f"-> {reference.shape[0]:,} genes x {reference.shape[1]} types")
    return reference


def read_follicle_markers(path):
    """The three follicle gene sets as {gene: quality score}, read until the first blank row.

    The paper scores every marker 1-5 in a `Score` column at the far right of each sheet:
    1 established marker, 2 well annotated and biologically significant, 3 well annotated
    but not biologically significant, 4 poorly annotated gene, 5 poor marker based on
    comparisons across clusters. Ignoring it is expensive — of the 19 theca markers on a
    5K panel, 6 are score 5 and 3 more are score 4, so more than half of what would be
    scored is flagged as bad by the people who assembled the list.

    The scores are carried into the asset rather than applied here, so the threshold
    stays a decision the notebook makes and revisiting it does not mean re-running this
    converter over the workbooks.
    """
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sets = {}
    for name, sheet in S5_SETS.items():
        worksheet = workbook[sheet]
        header = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        score_col = next(i for i, cell in enumerate(header)
                         if str(cell).strip() == "Score")
        genes = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            gene = row[0]
            # A blank ends the list; a footnote paragraph follows it in some sheets.
            if gene is None or not str(gene).strip():
                break
            try:
                score = int(row[score_col])
            except (TypeError, ValueError):
                # Unscored is not the same as badly scored, so it round-trips as null
                # and the notebook decides what to do with it.
                score = None
            genes[str(gene).strip()] = score
        sets[name] = genes
        by_score = {s: sum(1 for v in genes.values() if v == s) for s in (1, 2, 3, 4, 5)}
        print(f"  {name}: {len(genes)} genes  "
              + " ".join(f"score{s}={n}" for s, n in by_score.items() if n))
    workbook.close()
    return sets


def main():
    args = parse_args()
    supdir, outdir = Path(args.supdir), Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    print("Data S6 - major cell type mean expression:")
    reference = read_major_reference(supdir / "adm7506_Data_S6.xlsx")
    reference_path = outdir / "ovary_reference_major.csv.gz"
    reference.to_csv(reference_path)
    print(f"  wrote {reference_path}")

    print("\nData S7 - immune subtype mean expression:")
    immune = read_immune_reference(supdir / "adm7506_Data_S7.xlsx")
    immune_path = outdir / "ovary_reference_immune.csv.gz"
    immune.to_csv(immune_path)
    print(f"  wrote {immune_path}")

    print("\nData S5 - follicle marker sets:")
    sets = read_follicle_markers(supdir / "adm7506_Data_S5.xlsx")
    markers_path = outdir / "ovary_follicle_markers.yaml"
    markers_path.write_text(yaml.safe_dump(sets, sort_keys=True, default_flow_style=False))
    print(f"  wrote {markers_path}")

    if args.panel:
        import anndata as ad
        panel = set(ad.read_zarr(f"{args.panel.rstrip('/')}/tables/table").var_names)
        print(f"\nOverlap with a {len(panel):,}-gene panel:")
        for label, table in (("major reference ", reference), ("immune reference", immune)):
            shared = panel & set(table.index)
            print(f"  {label}: {len(shared):,} shared "
                  f"({len(shared) / len(panel) * 100:.1f}% of the panel)")
        for name, genes in sets.items():
            on_panel = panel & set(genes)
            established = {g for g, score in genes.items() if score == 1} & panel
            print(f"  {name:<10} {len(on_panel):>3} of {len(genes):>3} markers on panel, "
                  f"{len(established):>2} of them established (score 1)")


if __name__ == "__main__":
    main()
